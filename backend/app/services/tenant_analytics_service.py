from __future__ import annotations

import csv
import json
from dataclasses import dataclass, replace
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from backend.app.models import (
    Company,
    Dataset,
    DatasetRelationship,
    DatasetStatus,
    JobStatus,
    ModelRegistry,
    TrainingJob,
)
from backend.app.routers.datasets import _pipeline_status, _training_status
from backend.app.services.company_dataset_ingestion_service import CompanyDatasetIngestionService
from shared.ai_engine.contracts import TenantContext
from shared.ai_engine.dataset_ingestion.cleaning import CompanyDatasetCleaner
from shared.ai_engine.dataset_ingestion.prepared_dataset import PreparedCompanyDataset
from shared.ai_engine.dataset_ingestion.profiling import DatasetProfiler
from shared.ai_engine.dataset_ingestion.quality import assess_quality
from shared.ai_engine.dataset_ingestion.readiness import assess_capability_readiness


BUSINESS_METRIC_FIELDS = {
    "revenue": frozenset({"total_amount"}),
    "orders": frozenset({"order_id"}),
    "customers": frozenset({"customer_id"}),
    "average_order_value": frozenset({"total_amount", "order_id"}),
}

_ADDITIVE_FIELDS = frozenset(
    {"total_amount", "quantity", "unit_price", "inventory_level"}
)

# Large imports must never be fully reconstructed inside an API request. For
# business analytics we materialize a compact, tenant-scoped order rollup and
# cache it beside the source artifact. This keeps exact revenue/order KPIs while
# bounding memory for datasets with hundreds of thousands (or millions) of rows.
_DASHBOARD_MAX_PREPARED_ROWS = 50_000
_ROLLUP_VERSION = 1


@dataclass(frozen=True, slots=True)
class TenantAnalyticsSnapshot:
    company: Company | None
    datasets: tuple[Dataset, ...]
    statuses: tuple[str, ...]
    training_statuses: tuple[str, ...]
    prepared: tuple[PreparedCompanyDataset, ...]
    relationships: tuple[DatasetRelationship, ...]
    active_models: tuple[ModelRegistry, ...]
    capabilities: frozenset[str]
    status: str
    deferred_dataset_ids: frozenset[object] = frozenset()

    @property
    def currency(self) -> str:
        return self.company.currency_code if self.company is not None else "USD"

    def source_for(self, required_fields: frozenset[str]) -> PreparedCompanyDataset | None:
        candidates = [
            (item, self._with_derived_revenue(self._compose_from(item)))
            for item in self.prepared
        ]
        usable = [
            (anchor, composed)
            for anchor, composed in candidates
            if required_fields <= set(composed.canonical_columns.values())
        ]
        selected = max(
            usable,
            key=lambda pair: (
                len(required_fields & set(pair[0].canonical_columns.values())),
                sum(
                    required <= set(pair[1].canonical_columns.values())
                    for required in BUSINESS_METRIC_FIELDS.values()
                ),
                len(pair[1].canonical_columns),
                len(pair[1].rows),
            ),
            default=None,
        )
        return selected[1] if selected is not None else None

    @staticmethod
    def _with_derived_revenue(
        prepared: PreparedCompanyDataset,
    ) -> PreparedCompanyDataset:
        reverse = {
            canonical: original
            for original, canonical in prepared.canonical_columns.items()
        }
        if "total_amount" in reverse:
            return prepared
        quantity_column = reverse.get("quantity")
        price_column = reverse.get("unit_price")
        if quantity_column is None or price_column is None:
            return prepared

        derived_column = "__avenqo_total_amount"
        rows: list[dict[str, object]] = []
        for row in prepared.rows:
            output = dict(row)
            try:
                output[derived_column] = float(row[quantity_column]) * float(
                    row[price_column]
                )
            except (KeyError, TypeError, ValueError):
                output[derived_column] = None
            rows.append(output)
        return PreparedCompanyDataset(
            company_id=prepared.company_id,
            dataset_id=prepared.dataset_id,
            version=prepared.version,
            canonical_columns={
                **prepared.canonical_columns,
                derived_column: "total_amount",
            },
            rows=tuple(rows),
            profile=prepared.profile,
            mapping=prepared.mapping,
            cleaning_report=prepared.cleaning_report,
            quality=prepared.quality,
            capability_readiness=prepared.capability_readiness,
        )

    def _compose_from(self, anchor: PreparedCompanyDataset) -> PreparedCompanyDataset:
        prepared_by_id = {item.dataset_id: item for item in self.prepared}
        included = {anchor.dataset_id}
        rows = self._canonical_rows(anchor)
        fields = set(anchor.canonical_columns.values())
        enriched = False

        while True:
            best: tuple[DatasetRelationship, PreparedCompanyDataset, int] | None = None
            for relationship in self.relationships:
                if relationship.left_dataset_id in included:
                    other_id = relationship.right_dataset_id
                elif relationship.right_dataset_id in included:
                    other_id = relationship.left_dataset_id
                else:
                    continue
                if other_id in included or other_id not in prepared_by_id:
                    continue
                other = prepared_by_id[other_id]
                added = len(set(other.canonical_columns.values()) - fields)
                if added and (best is None or added > best[2]):
                    best = relationship, other, added
            if best is None:
                break

            relationship, other, _ = best
            other_rows = self._canonical_rows(other)
            join_field = relationship.canonical_field
            lookup: dict[str, dict[str, object]] = {}
            duplicate = False
            for row in other_rows:
                key = self._join_key(row.get(join_field))
                if key is None:
                    continue
                if key in lookup:
                    duplicate = True
                    break
                lookup[key] = row
            if duplicate or not lookup or join_field not in fields:
                included.add(other.dataset_id)
                continue

            join_keys = [
                key
                for row in rows
                if (key := self._join_key(row.get(join_field))) is not None
            ]
            added_fields = set(other.canonical_columns.values()) - fields
            if len(join_keys) != len(set(join_keys)):
                added_fields -= _ADDITIVE_FIELDS
            if not added_fields:
                included.add(other.dataset_id)
                continue

            rows = tuple(
                {
                    **row,
                    **{
                        field: value
                        for field, value in lookup.get(self._join_key(row.get(join_field)) or "", {}).items()
                        if field in added_fields
                    },
                }
                for row in rows
            )
            fields.update(added_fields)
            included.add(other.dataset_id)
            enriched = True

        if not enriched:
            return anchor
        return PreparedCompanyDataset(
            company_id=anchor.company_id,
            dataset_id=anchor.dataset_id,
            version=anchor.version,
            canonical_columns={field: field for field in fields},
            rows=rows,
            profile=anchor.profile,
            mapping=anchor.mapping,
            cleaning_report=anchor.cleaning_report,
            quality=anchor.quality,
            capability_readiness=anchor.capability_readiness,
        )

    @staticmethod
    def _canonical_rows(prepared: PreparedCompanyDataset) -> tuple[dict[str, object], ...]:
        reverse = {canonical: original for original, canonical in prepared.canonical_columns.items()}
        return tuple(
            {
                canonical: row[original]
                for canonical, original in reverse.items()
                if original in row
            }
            for row in prepared.rows
        )

    @staticmethod
    def _join_key(value: object | None) -> str | None:
        if value is None:
            return None
        normalized = str(value).strip().casefold()
        return normalized or None


class TenantAnalyticsService:
    """Loads shared, tenant-scoped analytics inputs without unbounded API memory."""

    def __init__(self, session: Session, ingestion: CompanyDatasetIngestionService) -> None:
        self._session = session
        self._ingestion = ingestion

    def load(self, tenant: TenantContext) -> TenantAnalyticsSnapshot:
        # All request-time business analytics use the same bounded path. Large
        # datasets are represented by an exact order-level rollup instead of
        # reloading/re-cleaning the entire source in the web process.
        return self._load(tenant, max_prepared_rows=_DASHBOARD_MAX_PREPARED_ROWS)

    def load_for_dashboard(self, tenant: TenantContext) -> TenantAnalyticsSnapshot:
        return self._load(tenant, max_prepared_rows=_DASHBOARD_MAX_PREPARED_ROWS)

    def _load(
        self,
        tenant: TenantContext,
        *,
        max_prepared_rows: int | None = None,
    ) -> TenantAnalyticsSnapshot:
        company = self._session.scalar(select(Company).where(Company.id == tenant.company_id))
        datasets = tuple(
            self._session.scalars(
                select(Dataset)
                .where(Dataset.company_id == tenant.company_id)
                .options(selectinload(Dataset.training_jobs))
                .order_by(Dataset.uploaded_at.desc())
            ).all()
        )
        statuses = tuple(_pipeline_status(dataset) for dataset in datasets)
        training_statuses = tuple(
            status
            for dataset in datasets
            if (status := _training_status(dataset)) not in {None, "not_applicable"}
        )
        prepared, deferred_dataset_ids = self._prepared_ready_datasets(
            tenant, datasets, max_prepared_rows=max_prepared_rows
        )
        prepared_ids = {item.dataset_id for item in prepared}
        relationships = tuple(
            self._session.scalars(
                select(DatasetRelationship).where(
                    DatasetRelationship.company_id == tenant.company_id,
                    DatasetRelationship.left_dataset_id.in_(prepared_ids),
                    DatasetRelationship.right_dataset_id.in_(prepared_ids),
                )
            ).all()
        ) if prepared_ids else ()
        active_models = tuple(
            self._session.scalars(
                select(ModelRegistry)
                .join(TrainingJob, TrainingJob.id == ModelRegistry.training_job_id)
                .where(
                    ModelRegistry.company_id == tenant.company_id,
                    ModelRegistry.is_active.is_(True),
                    TrainingJob.company_id == tenant.company_id,
                    TrainingJob.status == JobStatus.COMPLETED,
                    TrainingJob.dataset_id.in_(prepared_ids),
                )
            ).all()
        ) if prepared_ids else ()
        snapshot = TenantAnalyticsSnapshot(
            company=company,
            datasets=datasets,
            statuses=statuses,
            training_statuses=training_statuses,
            prepared=prepared,
            relationships=relationships,
            active_models=active_models,
            capabilities=frozenset(),
            status=self._status(datasets, statuses, prepared, deferred_dataset_ids),
            deferred_dataset_ids=frozenset(deferred_dataset_ids),
        )
        return replace(
            snapshot,
            capabilities=frozenset(self._capabilities(snapshot, active_models)),
        )

    def _prepared_ready_datasets(
        self,
        tenant: TenantContext,
        datasets: tuple[Dataset, ...],
        *,
        max_prepared_rows: int | None = None,
    ) -> tuple[tuple[PreparedCompanyDataset, ...], set[object]]:
        result: list[PreparedCompanyDataset] = []
        deferred_dataset_ids: set[object] = set()
        for dataset in datasets:
            if dataset.status != DatasetStatus.READY or dataset.mapping is None:
                continue
            if (
                max_prepared_rows is not None
                and int(dataset.rows_count or 0) > max_prepared_rows
            ):
                try:
                    rollup = self._load_or_build_large_retail_rollup(tenant, dataset)
                except Exception:
                    rollup = None
                if rollup is not None:
                    result.append(rollup)
                else:
                    deferred_dataset_ids.add(dataset.id)
                continue
            try:
                result.append(self._ingestion.get_prepared_dataset(tenant, dataset.id))
            except Exception:
                continue
        return tuple(result), deferred_dataset_ids

    def _load_or_build_large_retail_rollup(
        self,
        tenant: TenantContext,
        dataset: Dataset,
    ) -> PreparedCompanyDataset | None:
        """Return an exact order-level KPI rollup for a large retail source.

        The cache contains only tenant business aggregates, never raw rows from
        another tenant. Revenue is summed from mapped total_amount or from
        quantity * unit_price. Distinct orders/customers remain exact because
        one row is retained per real order_id.
        """
        if dataset.company_id != tenant.company_id or dataset.mapping is None:
            return None
        accepted = dict(dataset.mapping.mapping_json.get("accepted") or {})
        reverse = {canonical: original for original, canonical in accepted.items()}
        order_col = reverse.get("order_id")
        amount_col = reverse.get("total_amount")
        quantity_col = reverse.get("quantity")
        price_col = reverse.get("unit_price")
        if order_col is None or (amount_col is None and (quantity_col is None or price_col is None)):
            return None

        current_version = max((item.version_number for item in dataset.versions), default=1)
        raw_path = Path(str(dataset.source))
        cache_path = raw_path.parent / f"retail-kpi-rollup-v{_ROLLUP_VERSION}.json"
        cached_rows: list[dict[str, object]] | None = None
        if cache_path.is_file():
            try:
                payload = json.loads(cache_path.read_text(encoding="utf-8"))
                if (
                    isinstance(payload, dict)
                    and payload.get("version") == _ROLLUP_VERSION
                    and payload.get("dataset_id") == str(dataset.id)
                    and payload.get("dataset_version") == current_version
                    and isinstance(payload.get("rows"), list)
                ):
                    cached_rows = [dict(row) for row in payload["rows"] if isinstance(row, dict)]
            except (OSError, json.JSONDecodeError, TypeError, ValueError):
                cached_rows = None

        if cached_rows is None:
            if not raw_path.is_file():
                return None
            rollup: dict[str, dict[str, object]] = {}
            customer_col = reverse.get("customer_id")
            timestamp_col = reverse.get("order_timestamp")
            for raw in self._iter_source_rows(raw_path):
                order_value = raw.get(order_col)
                if order_value is None or not str(order_value).strip():
                    continue
                order_id = str(order_value).strip()
                entry = rollup.setdefault(
                    order_id,
                    {
                        "order_id": order_id,
                        "customer_id": None,
                        "order_timestamp": None,
                        "total_amount": 0.0,
                    },
                )
                amount = self._amount(raw, amount_col, quantity_col, price_col)
                if amount is not None:
                    entry["total_amount"] = float(entry["total_amount"] or 0.0) + amount
                if customer_col and entry["customer_id"] is None:
                    customer = raw.get(customer_col)
                    if customer is not None and str(customer).strip():
                        entry["customer_id"] = str(customer).strip()
                if timestamp_col and entry["order_timestamp"] is None:
                    timestamp = self._iso_datetime(raw.get(timestamp_col))
                    if timestamp is not None:
                        entry["order_timestamp"] = timestamp
            cached_rows = list(rollup.values())
            if not cached_rows:
                return None
            try:
                cache_path.write_text(
                    json.dumps(
                        {
                            "version": _ROLLUP_VERSION,
                            "dataset_id": str(dataset.id),
                            "dataset_version": current_version,
                            "rows": cached_rows,
                        },
                        ensure_ascii=False,
                        separators=(",", ":"),
                    ),
                    encoding="utf-8",
                )
            except OSError:
                # Cache persistence is an optimization only; computed rows are
                # still safe to use for this request.
                pass

        columns = ("order_id", "customer_id", "order_timestamp", "total_amount")
        cleaner = CompanyDatasetCleaner()
        canonical = {name: name for name in columns}
        cleaned_rows, cleaning_report = cleaner.clean(cached_rows, canonical)
        profile = DatasetProfiler().profile(cleaned_rows, columns)
        return PreparedCompanyDataset(
            company_id=tenant.company_id,
            dataset_id=dataset.id,
            version=current_version,
            canonical_columns=canonical,
            rows=tuple(cleaned_rows),
            profile=profile,
            mapping=(),
            cleaning_report=cleaning_report,
            quality=assess_quality(cleaning_report),
            capability_readiness=assess_capability_readiness(set(canonical.values())),
        )

    @staticmethod
    def _iter_source_rows(path: Path) -> Iterable[dict[str, object]]:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                yield from csv.DictReader(handle)
            return
        if suffix in {".xlsx", ".xlsm"}:
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                sheet = workbook.active
                rows = sheet.iter_rows(values_only=True)
                headers = next(rows, None)
                if not headers:
                    return
                names = [str(value).strip() if value is not None else "" for value in headers]
                for values in rows:
                    yield {
                        names[index]: value
                        for index, value in enumerate(values)
                        if index < len(names) and names[index]
                    }
            finally:
                workbook.close()
            return
        # Other formats keep the normal prepared path when small; large
        # unsupported formats are deferred instead of risking unbounded memory.
        return

    @staticmethod
    def _amount(
        row: dict[str, object],
        amount_col: str | None,
        quantity_col: str | None,
        price_col: str | None,
    ) -> float | None:
        try:
            if amount_col is not None and row.get(amount_col) not in {None, ""}:
                return float(row[amount_col])
            if quantity_col is not None and price_col is not None:
                quantity = row.get(quantity_col)
                price = row.get(price_col)
                if quantity not in {None, ""} and price not in {None, ""}:
                    return float(quantity) * float(price)
        except (TypeError, ValueError):
            return None
        return None

    @staticmethod
    def _iso_datetime(value: object | None) -> str | None:
        if isinstance(value, datetime):
            return value.isoformat()
        if isinstance(value, date):
            return datetime.combine(value, datetime.min.time()).isoformat()
        if isinstance(value, str) and value.strip():
            text = value.strip()
            try:
                return datetime.fromisoformat(text).isoformat()
            except ValueError:
                return text
        return None

    @staticmethod
    def _capabilities(
        snapshot: TenantAnalyticsSnapshot,
        active_models: tuple[ModelRegistry, ...],
    ) -> set[str]:
        capabilities = {model.task_code for model in active_models}
        capabilities.update(
            key
            for key, required in BUSINESS_METRIC_FIELDS.items()
            if snapshot.source_for(required) is not None
        )
        for dataset in snapshot.prepared:
            capabilities.update(
                readiness.capability
                for readiness in dataset.capability_readiness
                if readiness.ready
            )
        return capabilities

    @staticmethod
    def _status(
        datasets: tuple[Dataset, ...],
        statuses: tuple[str, ...],
        prepared: tuple[PreparedCompanyDataset, ...],
        deferred_dataset_ids: set[object],
    ) -> str:
        if not datasets:
            return "no_data"
        if (prepared or deferred_dataset_ids) and any(status != "ready" for status in statuses):
            return "partial_ready"
        if prepared or deferred_dataset_ids:
            return "ready"
        if statuses and all(status == "failed" for status in statuses):
            return "error"
        return "processing"
