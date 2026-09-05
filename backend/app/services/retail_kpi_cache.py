"""Fixed-size dashboard summaries, built outside the request path.

The ingestion path primes the cache; existing imports are queued lazily. Only
immutable source descriptors cross into the worker, never SQLAlchemy sessions.
One worker and a bounded queue limit concurrent work. Source rows are streamed;
exact distinct-ID sets live only in the worker, not in dashboard requests.
Atomic publication and identity checks prevent partial or stale cache reads.
Interrupted work is safely queued again after a process restart.
"""
from __future__ import annotations

import csv
import hashlib
import json
import logging
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from tempfile import NamedTemporaryFile
from threading import Lock

_pool = ThreadPoolExecutor(max_workers=1, thread_name_prefix="retail-kpi")
_lock = Lock()
_pending: set[str] = set()
_failed: set[str] = set()
_log = logging.getLogger(__name__)


def current_version(dataset):
    versions = [v for v in dataset.versions if v.is_current]
    return versions[0] if len(versions) == 1 else None


def descriptor(tenant, dataset):
    if dataset.company_id != tenant.company_id:
        return None
    version = current_version(dataset)
    if version is None or not version.artifact_path:
        return None
    path = Path(version.artifact_path)
    try:
        stat = path.stat()
        if not path.is_file():
            return None
    except OSError:
        return None
    return {
        "format": 2, "company": str(tenant.company_id), "dataset": str(dataset.id),
        "version": version.version_number, "version_id": version.id,
        "artifact": str(path.resolve()), "checksum": version.checksum,
        "size": stat.st_size, "mtime": stat.st_mtime_ns,
        "mapping": dict(dataset.mapping.mapping_json.get("accepted") or {}),
    }


def cache_path(spec):
    digest = hashlib.sha256(json.dumps(spec, sort_keys=True).encode()).hexdigest()
    return Path(spec["artifact"]).parent / f"retail-kpi-{digest}.json"


def read_or_schedule(tenant, dataset):
    spec = descriptor(tenant, dataset)
    if spec is None:
        return "SOURCE_UNAVAILABLE", None
    path = cache_path(spec)
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if (payload["source"] == spec
                and isinstance(payload.get("current"), dict)
                and isinstance(payload.get("previous"), dict)
                and isinstance(payload.get("period"), dict)
                and set(payload["period"]) == {"start", "end", "comparison_start", "comparison_end"}):
            return "AVAILABLE", payload
    except (OSError, ValueError, KeyError, TypeError):
        pass
    key = str(path)
    with _lock:
        if key in _failed:
            return "SOURCE_UNAVAILABLE", None
        if key not in _pending and len(_pending) < 16:
            _pending.add(key)
            try:
                _pool.submit(_run, spec, key)
            except RuntimeError:
                _pending.discard(key)
                return "SOURCE_UNAVAILABLE", None
    return "PROCESSING", None


def _run(spec, key):
    try:
        build(spec)
    except Exception:
        _log.exception("Retail KPI preparation failed for dataset %s", spec["dataset"])
        with _lock:
            _failed.add(key)
    finally:
        with _lock:
            _pending.discard(key)


def rows(path):
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as handle:
            yield from csv.DictReader(handle)
    elif path.suffix.lower() in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook
        book = load_workbook(path, read_only=True, data_only=True)
        try:
            stream = book.active.iter_rows(values_only=True)
            names = next(stream, ())
            for values in stream:
                yield dict(zip(names, values))
        finally:
            book.close()
    elif path.suffix.lower() == ".parquet":
        import pyarrow.parquet as pq
        for batch in pq.ParquetFile(path).iter_batches(batch_size=8192):
            yield from batch.to_pylist()
    else:
        raise ValueError("Unsupported large retail artifact format")


def timestamp(value):
    try:
        parsed = datetime.fromisoformat(str(value))
        return parsed.replace(tzinfo=timezone.utc) if parsed.tzinfo is None else parsed.astimezone(timezone.utc)
    except (ValueError, TypeError):
        return None


def build(spec):
    path = Path(spec["artifact"])
    reverse = {v: k for k, v in spec["mapping"].items()}
    date_col = reverse.get("order_timestamp")
    end = None
    if date_col:
        for row in rows(path):
            value = timestamp(row.get(date_col))
            if value is not None:
                end = max(end, value) if end else value
    start = end - timedelta(days=29) if end else None
    comparison_start = start - timedelta(days=30) if start else None
    comparison_end = start - timedelta(microseconds=1) if start else None
    buckets = [{"revenue": Decimal(0), "orders": set(), "customers": set(), "rows": 0} for _ in range(2)]
    for row in rows(path):
        index = 0
        if end:
            value = timestamp(row.get(date_col))
            if value is None:
                continue
            if start <= value <= end:
                index = 0
            elif comparison_start <= value <= comparison_end:
                index = 1
            else:
                continue
        bucket = buckets[index]
        bucket["rows"] += 1
        for field in ("orders", "customers"):
            column = reverse.get("order_id" if field == "orders" else "customer_id")
            value = row.get(column)
            if value is not None and str(value).strip():
                bucket[field].add(str(value).strip())
        try:
            value = row.get(reverse.get("total_amount"))
            if value is None or value == "":
                amount = Decimal(str(row.get(reverse.get("quantity")))) * Decimal(str(row.get(reverse.get("unit_price"))))
            else:
                amount = Decimal(str(value))
            if amount.is_finite():
                bucket["revenue"] += amount
        except (InvalidOperation, ValueError, TypeError):
            pass
    fields = set(reverse)
    revenue_ready = "total_amount" in fields or {"quantity", "unit_price"} <= fields
    available = {"revenue": revenue_ready, "orders": "order_id" in fields,
                 "customers": "customer_id" in fields, "average_order_value": revenue_ready and "order_id" in fields}
    def metrics(bucket):
        orders = len(bucket["orders"])
        values = {"revenue": float(round(bucket["revenue"], 2)), "orders": orders,
                  "customers": len(bucket["customers"]),
                  "average_order_value": float(round(bucket["revenue"] / orders, 2)) if orders else 0.0}
        return {key: value for key, value in values.items() if available[key]}
    payload = {"source": spec, "current": metrics(buckets[0]),
               "previous": metrics(buckets[1]) if buckets[1]["rows"] else {},
               "period": {key: value.isoformat() if value else None for key, value in
                          dict(start=start, end=end, comparison_start=comparison_start, comparison_end=comparison_end).items()}}
    stat = path.stat()
    if (stat.st_size, stat.st_mtime_ns) != (spec["size"], spec["mtime"]):
        raise ValueError("Artifact changed during KPI preparation")
    target = cache_path(spec)
    with NamedTemporaryFile(mode="w", encoding="utf-8", dir=target.parent, delete=False) as handle:
        temporary = Path(handle.name)
        json.dump(payload, handle, separators=(",", ":"))
    try:
        temporary.replace(target)
    finally:
        temporary.unlink(missing_ok=True)
    return payload
